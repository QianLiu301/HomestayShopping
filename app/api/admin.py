import os
from flask import request, current_app, Response
from app.api import api_bp
from app.models import (
    Admin, Product, Category, Vehicle, Location, Setting, Coupon,
    TransferOrder, ShopOrder, OrderItem,
    TicketAttraction, TicketPackage, TicketOrder, TicketTraveler,
    TicketVoucher, TicketTransportPrice
)
from app import db, bcrypt
from app.utils import (
    success_response, error_response, admin_required, paginate_query, escape_like
)
from app.utils.storage import upload_file, get_r2_file
from app.translations import auto_fill_translations


def _get_product_sales_map(product_ids):
    if not product_ids:
        return {}

    rows = db.session.query(
        OrderItem.product_id,
        db.func.coalesce(db.func.sum(OrderItem.quantity), 0).label('sales')
    ).join(ShopOrder, OrderItem.order_id == ShopOrder.id).filter(
        OrderItem.product_id.in_(product_ids),
        ShopOrder.status != 4,
        ShopOrder.payment_status == 1
    ).group_by(OrderItem.product_id).all()

    return {product_id: int(sales or 0) for product_id, sales in rows}


# ==================== 文件上传 ====================

def allowed_file(filename):
    allowed = current_app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'gif', 'webp'})
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed


@api_bp.route('/admin/upload', methods=['POST'])
@admin_required
def admin_upload_file():
    """上传文件"""
    if 'file' not in request.files:
        return error_response('没有选择文件')

    file = request.files['file']
    if file.filename == '':
        return error_response('没有选择文件')

    if not allowed_file(file.filename):
        return error_response('不支持的文件格式，请上传 png/jpg/jpeg/gif/webp')

    upload_folder = current_app.config['UPLOAD_FOLDER']
    try:
        url = upload_file(file, upload_folder)
        current_app.logger.info(f'文件上传成功: {url}')
        return success_response({'url': url}, '上传成功')
    except Exception as e:
        current_app.logger.error(f'文件上传失败: {e}')
        return error_response(f'上传失败: {str(e)}', 500)


# ==================== R2 图片代理 ====================

@api_bp.route('/images/<path:key>', methods=['GET'])
def proxy_r2_image(key):
    """代理 R2 图片，避免依赖 R2 公开访问 URL（带内存缓存 + 强浏览器缓存）"""
    import re
    if not re.match(r'^[a-f0-9]{32}\.\w+$', key):
        return error_response('Invalid key', 400)

    # 检查浏览器 ETag 缓存
    if_none_match = request.headers.get('If-None-Match')
    if if_none_match:
        # 图片内容不变，文件名即可作为 ETag
        etag = f'"{key}"'
        if if_none_match == etag:
            return Response(status=304)

    data, content_type = get_r2_file(key)
    if data is None:
        return error_response('Image not found', 404)

    # 确保 Content-Type 精确，避免 CORB
    if not content_type or content_type == 'application/octet-stream':
        ext = key.rsplit('.', 1)[-1].lower()
        content_type_map = {
            'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
            'png': 'image/png', 'gif': 'image/gif', 'webp': 'image/webp'
        }
        content_type = content_type_map.get(ext, 'image/jpeg')

    etag = f'"{key}"'
    return Response(data, content_type=content_type, headers={
        'Cache-Control': 'public, max-age=31536000, immutable',
        'ETag': etag,
        'X-Content-Type-Options': 'nosniff',  # 防止 MIME 嗅探，避免 CORB
        'Access-Control-Allow-Origin': '*',   # 明确允许跨域
    })


# ==================== 商品管理 ====================

@api_bp.route('/admin/products', methods=['GET'])
@admin_required
def admin_get_products():
    """获取商品列表（管理端）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    category_id = request.args.get('category_id', type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    
    query = Product.query
    
    if category_id:
        query = query.filter_by(category_id=category_id)
    if status is not None:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(
            (Product.name_zh.ilike(f'%{escape_like(keyword)}%')) |
            (Product.name_en.ilike(f'%{escape_like(keyword)}%'))
        )
    
    query = query.order_by(Product.sort_order.desc(), Product.id.desc())
    result = paginate_query(query, page, per_page)
    product_ids = [p.id for p in result['items']]
    sales_map = _get_product_sales_map(product_ids)

    for product in result['items']:
        product.sales_count = sales_map.get(product.id, 0)

    return success_response({
        'list': [p.to_dict() for p in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/products', methods=['POST'])
@admin_required
def admin_create_product():
    """创建商品"""
    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'desc'])

    name_zh = data.get('name_zh') or data.get('name_en')
    name_en = data.get('name_en') or data.get('name_zh')

    if not name_zh and not name_en:
        return error_response('商品名称不能为空')
    if data.get('price') is None:
        return error_response('价格不能为空')

    name_fields = {
        'name_zh': name_zh,
        'name_en': name_en,
        'name_ru': data.get('name_ru'),
        'name_es': data.get('name_es')
    }
    for field_name, field_value in name_fields.items():
        if field_value and len(str(field_value)) > 1000:
            return error_response(f'{field_name} 不可以超过1000个字符')

    product = Product(
        category_id=data.get('category_id'),
        name_zh=name_zh,
        name_en=name_en,
        name_ru=data.get('name_ru'),
        name_es=data.get('name_es'),
        desc_zh=data.get('desc_zh') or data.get('desc_en'),
        desc_en=data.get('desc_en') or data.get('desc_zh'),
        desc_ru=data.get('desc_ru'),
        desc_es=data.get('desc_es'),
        price=data.get('price'),
        original_price=data.get('original_price'),
        images=data.get('images', []),
        specs=data.get('specs'),
        sort_order=data.get('sort_order', 0),
        is_featured=data.get('is_featured', False),
        status=data.get('status', 1)
    )
    
    db.session.add(product)
    db.session.commit()
    
    return success_response(product.to_dict(), '创建成功')


@api_bp.route('/admin/products/<int:product_id>', methods=['PUT'])
@admin_required
def admin_update_product(product_id):
    """更新商品"""
    product = Product.query.get(product_id)
    if not product:
        return error_response('商品不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'desc'])

    name_fields = ['name_zh', 'name_en', 'name_ru', 'name_es']
    for field in name_fields:
        if field in data and data[field] and len(str(data[field])) > 1000:
            return error_response(f'{field} 不可以超过1000个字符')

    # 更新字段
    fields = [
        'category_id', 'name_zh', 'name_en', 'name_ru', 'name_es',
        'desc_zh', 'desc_en', 'desc_ru', 'desc_es',
        'price', 'original_price', 'images', 'specs',
        'sort_order', 'is_featured', 'status'
    ]
    
    for field in fields:
        if field in data:
            setattr(product, field, data[field])
    
    db.session.commit()
    
    return success_response(product.to_dict(), '更新成功')


@api_bp.route('/admin/products/<int:product_id>', methods=['DELETE'])
@admin_required
def admin_delete_product(product_id):
    """删除商品"""
    product = Product.query.get(product_id)
    if not product:
        return error_response('商品不存在', 404)

    if product.order_items:
        return error_response('该商品已有历史订单关联，无法删除，请改为下架处理', 400)

    db.session.delete(product)
    db.session.commit()

    return success_response(None, '删除成功')


# ==================== 分类管理 ====================

@api_bp.route('/admin/categories', methods=['GET'])
@admin_required
def admin_get_categories():
    """获取分类列表（管理端）"""
    categories = Category.query.order_by(Category.sort_order.desc()).all()
    return success_response([c.to_dict() for c in categories])


@api_bp.route('/admin/categories', methods=['POST'])
@admin_required
def admin_create_category():
    """创建分类"""
    data = request.get_json()
    data = auto_fill_translations(data, ['name'])

    name_zh = data.get('name_zh')
    name_en = data.get('name_en') or data.get('name_zh')

    if not name_zh:
        return error_response('请填写中文分类名称')

    category = Category(
        name_zh=name_zh,
        name_en=name_en,
        name_ru=data.get('name_ru'),
        name_es=data.get('name_es'),
        icon=data.get('icon'),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    
    db.session.add(category)
    db.session.commit()
    
    return success_response(category.to_dict(), '创建成功')


@api_bp.route('/admin/categories/<int:category_id>', methods=['PUT'])
@admin_required
def admin_update_category(category_id):
    """更新分类"""
    category = Category.query.get(category_id)
    if not category:
        return error_response('分类不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['name'])

    if 'name_zh' in data and not data.get('name_zh'):
        return error_response('请填写中文分类名称')

    fields = ['name_zh', 'name_en', 'name_ru', 'name_es', 'icon', 'sort_order', 'status']
    for field in fields:
        if field in data:
            setattr(category, field, data[field])
    
    db.session.commit()
    
    return success_response(category.to_dict(), '更新成功')


@api_bp.route('/admin/categories/<int:category_id>', methods=['DELETE'])
@admin_required
def admin_delete_category(category_id):
    """删除分类"""
    category = Category.query.get(category_id)
    if not category:
        return error_response('分类不存在', 404)
    
    # 检查是否有商品使用此分类
    if Product.query.filter_by(category_id=category_id).first():
        return error_response('该分类下有商品，无法删除')
    
    db.session.delete(category)
    db.session.commit()
    
    return success_response(None, '删除成功')


# ==================== 车型管理 ====================

@api_bp.route('/admin/vehicles', methods=['GET'])
@admin_required
def admin_get_vehicles():
    """获取车型列表（管理端）"""
    vehicles = Vehicle.query.order_by(Vehicle.sort_order.desc()).all()
    return success_response([v.to_dict() for v in vehicles])


@api_bp.route('/admin/vehicles', methods=['POST'])
@admin_required
def admin_create_vehicle():
    """创建车型"""
    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'desc', 'model', 'capacity_desc'])

    name_zh = data.get('name_zh') or data.get('name_en')
    name_en = data.get('name_en') or data.get('name_zh')

    if not name_zh and not name_en:
        return error_response('车型名称不能为空')

    luggage_28 = int(data.get('luggage_28') or 0)
    luggage_24 = int(data.get('luggage_24') or 0)
    luggage_capacity = int(data.get('luggage_capacity') or (luggage_28 + luggage_24) or 0)
    images = data.get('images') or ([] if not data.get('image') else [data.get('image')])
    if not isinstance(images, list):
        images = [images]
    images = [str(item).strip() for item in images if str(item).strip()]

    vehicle = Vehicle(
        name_zh=name_zh,
        name_en=name_en,
        name_ru=data.get('name_ru'),
        name_es=data.get('name_es'),
        desc_zh=data.get('desc_zh') or data.get('desc_en'),
        desc_en=data.get('desc_en') or data.get('desc_zh'),
        desc_ru=data.get('desc_ru'),
        desc_es=data.get('desc_es'),
        model_zh=data.get('model_zh') or data.get('model_en'),
        model_en=data.get('model_en') or data.get('model_zh'),
        model_ru=data.get('model_ru'),
        model_es=data.get('model_es'),
        seats=data.get('seats', 5),
        luggage_capacity=luggage_capacity,
        luggage_28=luggage_28,
        luggage_24=luggage_24,
        capacity_desc_zh=data.get('capacity_desc_zh') or data.get('capacity_desc_en'),
        capacity_desc_en=data.get('capacity_desc_en') or data.get('capacity_desc_zh'),
        capacity_desc_ru=data.get('capacity_desc_ru'),
        capacity_desc_es=data.get('capacity_desc_es'),
        extra_price=data.get('extra_price', 0),
        pickup_price=data.get('pickup_price', 0),
        dropoff_price=data.get('dropoff_price', 0),
        combo_price=data.get('combo_price', 0),
        image=images[0] if images else data.get('image'),
        images=images,
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    
    db.session.add(vehicle)
    db.session.commit()
    
    return success_response(vehicle.to_dict(), '创建成功')


@api_bp.route('/admin/vehicles/<int:vehicle_id>', methods=['PUT'])
@admin_required
def admin_update_vehicle(vehicle_id):
    """更新车型"""
    vehicle = Vehicle.query.get(vehicle_id)
    if not vehicle:
        return error_response('车型不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'desc', 'model', 'capacity_desc'])

    fields = [
        'name_zh', 'name_en', 'name_ru', 'name_es',
        'desc_zh', 'desc_en', 'desc_ru', 'desc_es',
        'model_zh', 'model_en', 'model_ru', 'model_es',
        'seats', 'luggage_capacity', 'luggage_28', 'luggage_24',
        'capacity_desc_zh', 'capacity_desc_en', 'capacity_desc_ru', 'capacity_desc_es',
        'extra_price', 'pickup_price', 'dropoff_price', 'combo_price', 'sort_order', 'status'
    ]

    for field in fields:
        if field in data:
            setattr(vehicle, field, data[field])

    if 'luggage_28' in data or 'luggage_24' in data:
        luggage_28 = vehicle.luggage_28 or 0
        luggage_24 = vehicle.luggage_24 or 0
        if 'luggage_capacity' not in data:
            vehicle.luggage_capacity = luggage_28 + luggage_24

    if 'images' in data or 'image' in data:
        images = data.get('images')
        if images is None:
            images = [] if not data.get('image') else [data.get('image')]
        if not isinstance(images, list):
            images = [images]
        images = [str(item).strip() for item in images if str(item).strip()]
        vehicle.images = images
        vehicle.image = images[0] if images else None
    
    db.session.commit()
    
    return success_response(vehicle.to_dict(), '更新成功')


@api_bp.route('/admin/vehicles/<int:vehicle_id>', methods=['DELETE'])
@admin_required
def admin_delete_vehicle(vehicle_id):
    """删除车型"""
    vehicle = Vehicle.query.get(vehicle_id)
    if not vehicle:
        return error_response('车型不存在', 404)
    
    db.session.delete(vehicle)
    db.session.commit()
    
    return success_response(None, '删除成功')


# ==================== 民宿点管理 ====================

@api_bp.route('/admin/locations', methods=['GET'])
@admin_required
def admin_get_locations():
    """获取民宿点列表（管理端）"""
    locations = Location.query.order_by(Location.sort_order.desc()).all()
    return success_response([loc.to_dict() for loc in locations])


@api_bp.route('/admin/locations', methods=['POST'])
@admin_required
def admin_create_location():
    """创建民宿点"""
    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'address'])

    name_zh = data.get('name_zh') or data.get('name_en')
    name_en = data.get('name_en') or data.get('name_zh')
    address_zh = data.get('address_zh') or data.get('address_en')
    address_en = data.get('address_en') or data.get('address_zh')

    if not name_zh and not name_en:
        return error_response('名称不能为空')
    if not address_zh and not address_en:
        return error_response('地址不能为空')

    location = Location(
        name_zh=name_zh,
        name_en=name_en,
        name_ru=data.get('name_ru'),
        name_es=data.get('name_es'),
        address_zh=address_zh,
        address_en=address_en,
        address_ru=data.get('address_ru'),
        address_es=data.get('address_es'),
        district=data.get('district'),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    
    db.session.add(location)
    db.session.commit()
    
    return success_response(location.to_dict(), '创建成功')


@api_bp.route('/admin/locations/<int:location_id>', methods=['PUT'])
@admin_required
def admin_update_location(location_id):
    """更新民宿点"""
    location = Location.query.get(location_id)
    if not location:
        return error_response('民宿点不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'address'])

    fields = [
        'name_zh', 'name_en', 'name_ru', 'name_es',
        'address_zh', 'address_en', 'address_ru', 'address_es',
        'district', 'sort_order', 'status'
    ]
    
    for field in fields:
        if field in data:
            setattr(location, field, data[field])
    
    db.session.commit()
    
    return success_response(location.to_dict(), '更新成功')


@api_bp.route('/admin/locations/<int:location_id>', methods=['DELETE'])
@admin_required
def admin_delete_location(location_id):
    """删除民宿点"""
    location = Location.query.get(location_id)
    if not location:
        return error_response('民宿点不存在', 404)
    
    db.session.delete(location)
    db.session.commit()
    
    return success_response(None, '删除成功')


# ==================== 数据分析 ====================

@api_bp.route('/admin/analytics/top-products', methods=['GET'])
@admin_required
def admin_top_products():
    """热销商品排行"""
    from app.models import OrderItem
    limit = request.args.get('limit', 10, type=int)

    # 统计已完成订单（非取消）的商品销量
    rows = db.session.query(
        OrderItem.product_id,
        OrderItem.product_name,
        db.func.sum(OrderItem.quantity).label('total_qty'),
        db.func.sum(OrderItem.subtotal).label('total_revenue')
    ).join(ShopOrder, OrderItem.order_id == ShopOrder.id).filter(
        ShopOrder.status != 4  # 排除已取消
    ).group_by(OrderItem.product_id, OrderItem.product_name).order_by(
        db.func.sum(OrderItem.quantity).desc()
    ).limit(limit).all()

    result = [{
        'product_id': r.product_id,
        'product_name': r.product_name,
        'total_qty': int(r.total_qty),
        'total_revenue': round(float(r.total_revenue), 2)
    } for r in rows]

    return success_response(result)


@api_bp.route('/admin/analytics', methods=['GET'])
@admin_required
def admin_analytics():
    """营业数据分析 — DB-agnostic (aggregate in Python)"""
    from datetime import datetime, timedelta
    from collections import defaultdict

    period = request.args.get('period', 'month')  # month / quarter / half_year

    now = datetime.now()
    if period == 'quarter':
        start = datetime(now.year, ((now.month - 1) // 3) * 3 + 1, 1)
    elif period == 'half_year':
        start = datetime(now.year, now.month, 1) - timedelta(days=180)
        start = datetime(start.year, start.month, 1)
    else:
        start = datetime(now.year, now.month, 1)

    shop_cancelled = 4    # 商城订单取消状态码
    transfer_cancelled = 3  # 接送订单取消状态码
    # 门票订单状态：0=待支付 1=已支付 2=已使用 3=已取消 4=已退款
    ticket_cancelled_set = (3, 4)

    # Fetch raw rows (only need created_at + total_price)
    shop_rows = db.session.query(
        ShopOrder.created_at, ShopOrder.total_price
    ).filter(
        ShopOrder.created_at >= start,
        ShopOrder.status != shop_cancelled
    ).all()

    transfer_rows = db.session.query(
        TransferOrder.created_at, TransferOrder.total_price
    ).filter(
        TransferOrder.created_at >= start,
        TransferOrder.status != transfer_cancelled
    ).all()

    # 门票订单（已含门票接送的合并金额）
    from app.models import TicketOrder
    ticket_rows = db.session.query(
        TicketOrder.created_at, TicketOrder.total_price
    ).filter(
        TicketOrder.created_at >= start,
        ~TicketOrder.status.in_(ticket_cancelled_set)
    ).all()

    shop_count = len(shop_rows)
    shop_total = sum(float(r.total_price or 0) for r in shop_rows)
    transfer_count = len(transfer_rows)
    transfer_total = sum(float(r.total_price or 0) for r in transfer_rows)
    ticket_count = len(ticket_rows)
    ticket_total = sum(float(r.total_price or 0) for r in ticket_rows)

    # Build chart data in Python
    if period == 'month':
        # 按天
        fmt = '%Y-%m-%d'
        days = []
        d = start
        while d <= now:
            days.append(d.strftime(fmt))
            d += timedelta(days=1)
        labels = days
    else:
        # 按月
        fmt = '%Y-%m'
        labels = []
        d = start
        while d <= now:
            m = d.strftime(fmt)
            if m not in labels:
                labels.append(m)
            if d.month == 12:
                d = datetime(d.year + 1, 1, 1)
            else:
                d = datetime(d.year, d.month + 1, 1)

    shop_map = defaultdict(float)
    for r in shop_rows:
        if r.created_at:
            shop_map[r.created_at.strftime(fmt)] += float(r.total_price or 0)

    transfer_map = defaultdict(float)
    for r in transfer_rows:
        if r.created_at:
            transfer_map[r.created_at.strftime(fmt)] += float(r.total_price or 0)

    ticket_map = defaultdict(float)
    for r in ticket_rows:
        if r.created_at:
            ticket_map[r.created_at.strftime(fmt)] += float(r.total_price or 0)

    chart = {
        'labels': labels,
        'shop': [round(shop_map.get(l, 0), 2) for l in labels],
        'transfer': [round(transfer_map.get(l, 0), 2) for l in labels],
        'ticket': [round(ticket_map.get(l, 0), 2) for l in labels],
    }

    # 接送专员只能看到接送数据：其他业务清零隐藏
    from app.utils.permissions import current_role, ROLE_TRANSFER_OPS
    if current_role() == ROLE_TRANSFER_OPS:
        return success_response({
            'shop_count': 0, 'shop_total': 0,
            'transfer_count': transfer_count,
            'transfer_total': round(transfer_total, 2),
            'ticket_count': 0, 'ticket_total': 0,
            'total_revenue': round(transfer_total, 2),
            'chart': {
                'labels': chart['labels'],
                'shop': [0 for _ in chart['labels']],
                'transfer': chart['transfer'],
                'ticket': [0 for _ in chart['labels']],
            },
            'restricted_to_transfer': True,  # 前端据此隐藏其他卡片
        })

    return success_response({
        'shop_count': shop_count,
        'shop_total': round(shop_total, 2),
        'transfer_count': transfer_count,
        'transfer_total': round(transfer_total, 2),
        'ticket_count': ticket_count,
        'ticket_total': round(ticket_total, 2),
        'total_revenue': round(shop_total + transfer_total + ticket_total, 2),
        'chart': chart
    })


# ==================== 营业数据 Excel 导出 ====================

@api_bp.route('/admin/analytics/export', methods=['GET'])
@admin_required
def admin_analytics_export():
    """导出指定时间段的营业明细 Excel（财务对账用）

    生成 5 个 sheet：营收汇总 / 商城订单 / 接送订单 / 门票订单 / 退款记录
    接送专员只会拿到接送相关的两个 sheet。
    """
    from datetime import datetime, timedelta
    from io import BytesIO
    from flask import send_file
    from app.models import TicketOrder, ServiceReview  # noqa: F401
    from app.utils.permissions import current_role, ROLE_TRANSFER_OPS

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error_response('服务器缺少 openpyxl 依赖，请联系运维', 500)

    period = request.args.get('period', 'month')
    # 仅导出已支付订单（财务对账常用）
    paid_only = request.args.get('paid_only', '0') in ('1', 'true', 'True')
    is_transfer_only = current_role() == ROLE_TRANSFER_OPS

    # 时间窗口
    # 支持的 period 值：
    #   month         本月（当月 1 号 → 现在）—— 仪表盘视角
    #   quarter       本季度（本季 1 号 → 现在）
    #   half_year     近半年
    #   last_month    上个月（上月 1 号 → 上月最后一天）—— 财务对账首选
    #   last_3_months 前三个月（3 个月前 1 号 → 上月最后一天）—— 季度对账
    now = datetime.now()

    def _month_start(y, m):
        return datetime(y, m, 1)

    def _next_month_start(y, m):
        if m == 12:
            return datetime(y + 1, 1, 1)
        return datetime(y, m + 1, 1)

    end = now  # 默认结束于此刻

    if period == 'quarter':
        start = datetime(now.year, ((now.month - 1) // 3) * 3 + 1, 1)
        period_label = f'{start.year}年Q{(start.month - 1) // 3 + 1}'
    elif period == 'half_year':
        start = _month_start(now.year, now.month) - timedelta(days=180)
        start = _month_start(start.year, start.month)
        period_label = f'近半年({start.strftime("%Y-%m")} ~ {now.strftime("%Y-%m")})'
    elif period == 'last_month':
        # 上个月 1 号 → 上月最后一天 23:59:59
        last_month_end = _month_start(now.year, now.month) - timedelta(seconds=1)
        start = _month_start(last_month_end.year, last_month_end.month)
        end = last_month_end
        period_label = f'{start.strftime("%Y年%m月")}（完整月）'
    elif period == 'last_3_months':
        # 3 个月前 1 号 → 上月最后一天
        cur_first = _month_start(now.year, now.month)
        last_month_end = cur_first - timedelta(seconds=1)
        # 往前数 3 个完整月
        y, m = last_month_end.year, last_month_end.month - 2
        while m <= 0:
            m += 12
            y -= 1
        start = _month_start(y, m)
        end = last_month_end
        period_label = f'前三个月({start.strftime("%Y-%m")} ~ {last_month_end.strftime("%Y-%m")})'
    else:
        # 默认 month：本月
        start = _month_start(now.year, now.month)
        period_label = f'{start.strftime("%Y年%m月")}（截至{now.strftime("%m月%d日")}）'

    # ---- 样式 ----
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4A3728')
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='DDD0BC')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def write_headers(ws, headers):
        for col, label in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def status_label_shop(s):
        return {0: '待处理', 1: '已确认', 2: '配送中', 3: '已完成', 4: '已取消'}.get(s, str(s))

    def status_label_transfer(s):
        return {0: '待支付', 1: '已确认', 2: '已完成', 3: '已取消'}.get(s, str(s))

    def status_label_ticket(s):
        return {0: '待支付', 1: '已支付', 2: '已使用', 3: '已取消', 4: '已退款'}.get(s, str(s))

    def service_type_label(s):
        return {'pickup': '接机', 'dropoff': '送机', 'combo': '接送组合'}.get(s, s or '')

    def fmt_dt(dt):
        return dt.strftime('%Y-%m-%d %H:%M') if dt else ''

    # ============ 创建工作簿 ============
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 删默认 sheet

    # 拉数据（paid_only 时只取 payment_status=1 的）
    def _maybe_paid(q, model):
        if paid_only:
            q = q.filter(model.payment_status == 1)
        return q

    shop_rows = [] if is_transfer_only else _maybe_paid(
        ShopOrder.query.filter(ShopOrder.created_at >= start, ShopOrder.created_at <= end),
        ShopOrder
    ).order_by(ShopOrder.created_at.desc()).all()

    transfer_rows = _maybe_paid(
        TransferOrder.query.filter(TransferOrder.created_at >= start, TransferOrder.created_at <= end),
        TransferOrder
    ).order_by(TransferOrder.created_at.desc()).all()

    ticket_rows = [] if is_transfer_only else _maybe_paid(
        TicketOrder.query.filter(TicketOrder.created_at >= start, TicketOrder.created_at <= end),
        TicketOrder
    ).order_by(TicketOrder.created_at.desc()).all()

    # ========== Sheet 1: 营收汇总 ==========
    ws = wb.create_sheet('营收汇总')
    ws.merge_cells('A1:B1')
    scope_tag = ' [仅已支付]' if paid_only else ''
    title_cell = ws.cell(row=1, column=1, value=f'营业数据汇总 — {period_label}{scope_tag}')
    title_cell.font = Font(bold=True, size=14, color='4A3728')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value='导出时间')
    ws.cell(row=2, column=2, value=fmt_dt(now))
    ws.cell(row=3, column=1, value='数据范围')
    ws.cell(row=3, column=2, value='仅已支付订单' if paid_only else '全部订单（含未付/取消/退款）')

    shop_paid = [r for r in shop_rows if r.status != 4]
    transfer_paid = [r for r in transfer_rows if r.status != 3]
    ticket_paid = [r for r in ticket_rows if r.status not in (3, 4)]

    rows_data = [
        ('', ''),
        ('指标', '数值'),
    ]
    if not is_transfer_only:
        rows_data += [
            ('商城订单数', len(shop_paid)),
            ('商城营收 (¥)', round(sum(float(r.total_price or 0) for r in shop_paid), 2)),
        ]
    rows_data += [
        ('接送订单数', len(transfer_paid)),
        ('接送营收 (¥)', round(sum(float(r.total_price or 0) for r in transfer_paid), 2)),
    ]
    if not is_transfer_only:
        rows_data += [
            ('门票订单数', len(ticket_paid)),
            ('门票营收 (¥)', round(sum(float(r.total_price or 0) for r in ticket_paid), 2)),
            ('', ''),
            ('总订单数', len(shop_paid) + len(transfer_paid) + len(ticket_paid)),
            ('总营收 (¥)', round(
                sum(float(r.total_price or 0) for r in shop_paid)
                + sum(float(r.total_price or 0) for r in transfer_paid)
                + sum(float(r.total_price or 0) for r in ticket_paid), 2
            )),
        ]

    for i, (k, v) in enumerate(rows_data, start=4):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        if k in ('指标',):
            c1.font = header_font; c1.fill = header_fill; c1.alignment = center
            c2.font = header_font; c2.fill = header_fill; c2.alignment = center
        elif k.startswith('总'):
            c1.font = Font(bold=True, color='B98745')
            c2.font = Font(bold=True, color='B98745')
    autosize(ws, [22, 22])

    # ========== Sheet: 商城订单明细 ==========
    if not is_transfer_only:
        ws = wb.create_sheet('商城订单明细')
        headers = ['订单号', '客户', '电话', '邮箱', '金额(¥)', '订单状态', '支付状态', '退款状态', '下单时间']
        write_headers(ws, headers)
        for i, o in enumerate(shop_rows, start=2):
            ws.cell(row=i, column=1, value=o.order_no)
            ws.cell(row=i, column=2, value=o.contact_name)
            ws.cell(row=i, column=3, value=o.contact_phone or '')
            ws.cell(row=i, column=4, value=o.contact_email or '')
            ws.cell(row=i, column=5, value=float(o.total_price or 0))
            ws.cell(row=i, column=6, value=status_label_shop(o.status))
            ws.cell(row=i, column=7, value='已支付' if o.payment_status == 1 else '未支付')
            ws.cell(row=i, column=8, value='已退款' if o.refund_status == 1 else '—')
            ws.cell(row=i, column=9, value=fmt_dt(o.created_at))
        autosize(ws, [26, 14, 16, 24, 12, 12, 12, 12, 20])

    # ========== Sheet: 接送订单明细 ==========
    ws = wb.create_sheet('接送订单明细')
    headers = ['订单号', '客户', '电话', '邮箱', '交通方式', '服务类型', '金额(¥)', '订单状态', '支付状态', '退款状态', '下单时间']
    write_headers(ws, headers)
    for i, o in enumerate(transfer_rows, start=2):
        ws.cell(row=i, column=1, value=o.order_no)
        ws.cell(row=i, column=2, value=o.contact_name)
        ws.cell(row=i, column=3, value=o.contact_phone or '')
        ws.cell(row=i, column=4, value=o.contact_email or '')
        ws.cell(row=i, column=5, value='火车站' if (o.transport_mode or 'flight') == 'train' else '飞机')
        ws.cell(row=i, column=6, value=service_type_label(o.service_type))
        ws.cell(row=i, column=7, value=float(o.total_price or 0))
        ws.cell(row=i, column=8, value=status_label_transfer(o.status))
        ws.cell(row=i, column=9, value='已支付' if o.payment_status == 1 else '未支付')
        ws.cell(row=i, column=10, value='已退款' if o.refund_status == 1 else '—')
        ws.cell(row=i, column=11, value=fmt_dt(o.created_at))
    autosize(ws, [26, 14, 16, 24, 10, 12, 12, 12, 12, 12, 20])

    # ========== Sheet: 门票订单明细 ==========
    if not is_transfer_only:
        ws = wb.create_sheet('门票订单明细')
        headers = ['订单号', '客户', '电话', '邮箱', '景点ID', '游玩日期', '金额(¥)', '订单状态', '支付状态', '下单时间']
        write_headers(ws, headers)
        for i, o in enumerate(ticket_rows, start=2):
            ws.cell(row=i, column=1, value=o.order_no)
            ws.cell(row=i, column=2, value=o.contact_name)
            ws.cell(row=i, column=3, value=o.contact_phone or '')
            ws.cell(row=i, column=4, value=o.contact_email or '')
            ws.cell(row=i, column=5, value=o.attraction_id)
            ws.cell(row=i, column=6, value=str(o.visit_date) if o.visit_date else '')
            ws.cell(row=i, column=7, value=float(o.total_price or 0))
            ws.cell(row=i, column=8, value=status_label_ticket(o.status))
            ws.cell(row=i, column=9, value='已支付' if o.payment_status == 1 else '未支付')
            ws.cell(row=i, column=10, value=fmt_dt(o.created_at))
        autosize(ws, [26, 14, 16, 24, 10, 14, 12, 12, 12, 20])

    # ========== Sheet: 退款记录 ==========
    ws = wb.create_sheet('退款记录')
    headers = ['业务', '订单号', '客户', '金额(¥)', '退款金额(¥)', '订单状态', '退款时间', '取消时间']
    write_headers(ws, headers)
    row_idx = 2
    if not is_transfer_only:
        for o in shop_rows:
            if o.refund_status == 1:
                ws.cell(row=row_idx, column=1, value='商城')
                ws.cell(row=row_idx, column=2, value=o.order_no)
                ws.cell(row=row_idx, column=3, value=o.contact_name)
                ws.cell(row=row_idx, column=4, value=float(o.total_price or 0))
                ws.cell(row=row_idx, column=5, value=float(o.refund_amount or 0))
                ws.cell(row=row_idx, column=6, value=status_label_shop(o.status))
                ws.cell(row=row_idx, column=7, value=fmt_dt(o.refund_time))
                ws.cell(row=row_idx, column=8, value=fmt_dt(o.cancelled_at))
                row_idx += 1
    for o in transfer_rows:
        if o.refund_status == 1:
            ws.cell(row=row_idx, column=1, value='接送')
            ws.cell(row=row_idx, column=2, value=o.order_no)
            ws.cell(row=row_idx, column=3, value=o.contact_name)
            ws.cell(row=row_idx, column=4, value=float(o.total_price or 0))
            ws.cell(row=row_idx, column=5, value=float(o.refund_amount or 0))
            ws.cell(row=row_idx, column=6, value=status_label_transfer(o.status))
            ws.cell(row=row_idx, column=7, value=fmt_dt(o.refund_time))
            ws.cell(row=row_idx, column=8, value=fmt_dt(o.cancelled_at))
            row_idx += 1
    autosize(ws, [10, 26, 14, 12, 14, 12, 20, 20])

    # ============ 输出文件 ============
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    role_tag = '接送' if is_transfer_only else '全业务'
    paid_tag = '_已支付' if paid_only else ''
    filename = f'营业明细_{role_tag}{paid_tag}_{period}_{now.strftime("%Y%m%d_%H%M")}.xlsx'

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


# ==================== 订单管理 ====================

@api_bp.route('/admin/orders/shop', methods=['GET'])
@admin_required
def admin_get_shop_orders():
    """获取商城订单列表"""
    from datetime import datetime
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    query = ShopOrder.query

    if status is not None:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(
            (ShopOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.booking_no.ilike(f'%{escape_like(keyword)}%'))
        )
    if date_start:
        try:
            query = query.filter(ShopOrder.created_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            query = query.filter(ShopOrder.created_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
        except ValueError:
            pass

    query = query.order_by(ShopOrder.created_at.desc())
    result = paginate_query(query, page, per_page)

    return success_response({
        'list': [o.to_dict() for o in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/orders/shop/<int:order_id>', methods=['PUT'])
@admin_required
def admin_update_shop_order(order_id):
    """更新商城订单状态"""
    order = ShopOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    data = request.get_json()

    if 'status' in data:
        from app.models import china_now
        new_status = data['status']
        old_status = order.status
        order.status = new_status
        now = china_now()
        # 自动记录配送/完成时间
        if new_status == 2 and old_status != 2 and not order.delivery_time:
            order.delivery_time = now
        if new_status == 3 and old_status != 3 and not order.completed_time:
            order.completed_time = now
        if new_status == 4 and old_status != 4 and not order.cancelled_at:
            order.cancelled_at = now
            # 只有已支付的订单才需要退款
            if order.payment_status == 1 and order.refund_status == 0:
                order.refund_status = 1  # 待退款
                order.refund_amount = order.total_price
            elif order.payment_status == 0:
                # 未支付订单取消，无需退款
                order.refund_status = 0
                order.refund_amount = 0
    if 'remark' in data:
        order.remark = data['remark']
    if 'booking_no' in data:
        order.booking_no = data['booking_no']
    if 'resolved_address' in data:
        order.resolved_address = data['resolved_address']
    if 'checkout_date' in data:
        from datetime import date as date_cls
        raw = data['checkout_date']
        if raw:
            try:
                order.checkout_date = date_cls.fromisoformat(raw)
            except (ValueError, TypeError):
                pass
        else:
            order.checkout_date = None

    db.session.commit()

    return success_response(order.to_dict(), '更新成功')


@api_bp.route('/admin/orders/transfer', methods=['GET'])
@admin_required
def admin_get_transfer_orders():
    """获取接送机订单列表"""
    from datetime import datetime
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status = request.args.get('status', type=int)
    transport_mode = request.args.get('transport_mode', '')
    keyword = request.args.get('keyword', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    query = TransferOrder.query

    if status is not None:
        query = query.filter_by(status=status)
    if transport_mode in ('flight', 'train'):
        query = query.filter_by(transport_mode=transport_mode)
    if keyword:
        query = query.filter(
            (TransferOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_name.ilike(f'%{escape_like(keyword)}%'))
        )
    if date_start:
        try:
            query = query.filter(TransferOrder.created_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            query = query.filter(TransferOrder.created_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
        except ValueError:
            pass

    query = query.order_by(TransferOrder.created_at.desc())
    result = paginate_query(query, page, per_page)

    return success_response({
        'list': [o.to_dict() for o in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/orders/transfer/<int:order_id>', methods=['PUT'])
@admin_required
def admin_update_transfer_order(order_id):
    """更新接送机订单状态"""
    order = TransferOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)
    
    data = request.get_json()

    if 'status' in data:
        from app.models import china_now
        new_status = data['status']
        old_status = order.status
        order.status = new_status
        if new_status == 3 and old_status != 3 and not order.cancelled_at:
            order.cancelled_at = china_now()
            # 只有已支付的订单才需要退款
            if order.payment_status == 1 and order.refund_status == 0:
                order.refund_status = 1  # 待退款
                order.refund_amount = order.total_price
            elif order.payment_status == 0:
                # 未支付订单取消，无需退款
                order.refund_status = 0
                order.refund_amount = 0
    if 'remark' in data:
        order.remark = data['remark']
    if 'booking_no' in data:
        order.booking_no = data['booking_no']
    if 'resolved_address' in data:
        order.resolved_address = data['resolved_address']

    db.session.commit()

    return success_response(order.to_dict(), '更新成功')


# ==================== 批量删除订单 ====================

@api_bp.route('/admin/orders/shop/batch-delete', methods=['POST'])
@admin_required
def admin_batch_delete_shop_orders():
    """批量删除商城订单"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return error_response('请选择要删除的订单')
    orders = ShopOrder.query.filter(ShopOrder.id.in_(ids)).all()
    count = len(orders)
    for o in orders:
        db.session.delete(o)
    db.session.commit()
    return success_response({'deleted': count}, f'已删除 {count} 个订单')


@api_bp.route('/admin/orders/transfer/batch-delete', methods=['POST'])
@admin_required
def admin_batch_delete_transfer_orders():
    """批量删除接送机订单"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return error_response('请选择要删除的订单')
    orders = TransferOrder.query.filter(TransferOrder.id.in_(ids)).all()
    count = len(orders)
    for o in orders:
        db.session.delete(o)
    db.session.commit()
    return success_response({'deleted': count}, f'已删除 {count} 个订单')


@api_bp.route('/admin/ticket-orders/batch-delete', methods=['POST'])
@admin_required
def admin_batch_delete_ticket_orders():
    """批量删除门票订单"""
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return error_response('请选择要删除的订单')
    orders = TicketOrder.query.filter(TicketOrder.id.in_(ids)).all()
    count = len(orders)
    for o in orders:
        db.session.delete(o)
    db.session.commit()
    return success_response({'deleted': count}, f'已删除 {count} 个订单')


# ==================== 确认收款 ====================

@api_bp.route('/admin/orders/shop/<int:order_id>/confirm-payment', methods=['POST'])
@admin_required
def admin_confirm_shop_payment(order_id):
    """管理员确认商城订单已收款"""
    from app.models import china_now
    order = ShopOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)
    order.payment_status = 1
    order.payment_time = china_now()
    db.session.commit()
    return success_response(order.to_dict(), '已确认收款')


@api_bp.route('/admin/orders/transfer/<int:order_id>/confirm-payment', methods=['POST'])
@admin_required
def admin_confirm_transfer_payment(order_id):
    """管理员确认接送机订单已收款"""
    from app.models import china_now
    order = TransferOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)
    order.payment_status = 1
    order.payment_time = china_now()
    db.session.commit()
    return success_response(order.to_dict(), '已确认收款')


# ==================== 配送管理 ====================

@api_bp.route('/admin/delivery/orders', methods=['GET'])
@admin_required
def admin_get_delivery_orders():
    """获取待配送/配送中的订单列表（分页，按期望送达/退房日期排序）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', type=int)  # 1=已确认(待配送), 2=配送中
    keyword = request.args.get('keyword', '')

    query = ShopOrder.query.filter(ShopOrder.status.in_([1, 2]))

    if status is not None:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(
            (ShopOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.booking_no.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.resolved_address.ilike(f'%{escape_like(keyword)}%'))
        )

    # 按期望送达日期优先，退房日期次之，最后创建时间
    query = query.order_by(
        ShopOrder.expected_delivery_date.asc().nullslast(),
        ShopOrder.checkout_date.asc().nullslast(),
        ShopOrder.created_at.asc()
    )

    result = paginate_query(query, page, per_page)

    orders = [o.to_dict() for o in result['items']]

    return success_response({
        'orders': orders,
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/delivery/start', methods=['POST'])
@admin_required
def admin_start_delivery():
    """批量开始配送"""
    from app.models import china_now
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return error_response('请选择要配送的订单')

    now = china_now()
    count = 0
    orders = ShopOrder.query.filter(ShopOrder.id.in_(ids), ShopOrder.status == 1).all()
    for o in orders:
        o.status = 2
        o.delivery_time = now
        count += 1

    db.session.commit()
    return success_response({'updated': count}, f'已开始配送 {count} 个订单')


@api_bp.route('/admin/delivery/complete', methods=['POST'])
@admin_required
def admin_complete_delivery():
    """批量确认送达"""
    from app.models import china_now
    data = request.get_json()
    ids = data.get('ids', [])
    if not ids:
        return error_response('请选择要完成的订单')

    now = china_now()
    count = 0
    orders = ShopOrder.query.filter(ShopOrder.id.in_(ids), ShopOrder.status == 2).all()
    for o in orders:
        o.status = 3
        o.completed_time = now
        count += 1

    db.session.commit()
    return success_response({'updated': count}, f'已完成配送 {count} 个订单')


# ==================== 优惠券管理 ====================

@api_bp.route('/admin/coupons', methods=['GET'])
@admin_required
def admin_get_coupons():
    """获取优惠券列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    
    query = Coupon.query.order_by(Coupon.created_at.desc())
    result = paginate_query(query, page, per_page)
    
    return success_response({
        'list': [c.to_dict() for c in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/coupons', methods=['POST'])
@admin_required
def admin_create_coupon():
    """创建优惠券"""
    data = request.get_json()
    
    if not data.get('code') or not data.get('name_zh'):
        return error_response('优惠券码和名称不能为空')
    
    # 检查优惠券码是否存在
    if Coupon.query.filter_by(code=data.get('code').upper()).first():
        return error_response('优惠券码已存在')
    
    coupon = Coupon(
        code=data.get('code').upper(),
        name_zh=data.get('name_zh'),
        name_en=data.get('name_en'),
        discount_type=data.get('discount_type', 'fixed'),
        discount_value=data.get('discount_value', 0),
        min_amount=data.get('min_amount', 0),
        max_discount=data.get('max_discount'),
        apply_to=data.get('apply_to', 'all'),
        total_count=data.get('total_count'),
        per_limit=data.get('per_limit', 1),
        start_time=data.get('start_time'),
        end_time=data.get('end_time'),
        status=data.get('status', 1)
    )
    
    db.session.add(coupon)
    db.session.commit()
    
    return success_response(coupon.to_dict(), '创建成功')


@api_bp.route('/admin/coupons/<int:coupon_id>', methods=['PUT'])
@admin_required
def admin_update_coupon(coupon_id):
    """更新优惠券"""
    coupon = Coupon.query.get(coupon_id)
    if not coupon:
        return error_response('优惠券不存在', 404)
    
    data = request.get_json()
    
    fields = [
        'name_zh', 'name_en', 'discount_type', 'discount_value',
        'min_amount', 'max_discount', 'apply_to', 'total_count',
        'per_limit', 'start_time', 'end_time', 'status'
    ]
    
    for field in fields:
        if field in data:
            setattr(coupon, field, data[field])
    
    db.session.commit()
    
    return success_response(coupon.to_dict(), '更新成功')


@api_bp.route('/admin/coupons/<int:coupon_id>', methods=['DELETE'])
@admin_required
def admin_delete_coupon(coupon_id):
    """删除优惠券"""
    coupon = Coupon.query.get(coupon_id)
    if not coupon:
        return error_response('优惠券不存在', 404)
    
    db.session.delete(coupon)
    db.session.commit()
    
    return success_response(None, '删除成功')


# ==================== 已取消订单管理（退款处理） ====================

@api_bp.route('/admin/orders/cancelled', methods=['GET'])
@admin_required
def admin_get_cancelled_orders():
    """获取已取消订单列表（商城+接送机统一查询）"""
    from datetime import datetime, timedelta
    from app.models import china_now
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 15, type=int)
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    refund_status = request.args.get('refund_status', type=int)
    keyword = request.args.get('keyword', '')
    
    # 接送专员只能看接送：完全跳过商城查询
    from app.utils.permissions import current_role, ROLE_TRANSFER_OPS
    transfer_ops_only = current_role() == ROLE_TRANSFER_OPS

    # 不设置默认日期筛选；未选择日期时返回全部已取消订单

    # 查询商城已取消订单 (status=4) — 接送专员看不到
    shop_query = ShopOrder.query.filter_by(status=4) if not transfer_ops_only else None
    if shop_query is not None:
        if date_start:
            try:
                shop_query = shop_query.filter(ShopOrder.cancelled_at >= datetime.fromisoformat(date_start))
            except ValueError:
                pass
        if date_end:
            try:
                shop_query = shop_query.filter(ShopOrder.cancelled_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
            except ValueError:
                pass
        if refund_status is not None:
            shop_query = shop_query.filter_by(refund_status=refund_status)
        if keyword:
            shop_query = shop_query.filter(
                (ShopOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
                (ShopOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
                (ShopOrder.contact_phone.ilike(f'%{escape_like(keyword)}%')) |
                (ShopOrder.contact_email.ilike(f'%{escape_like(keyword)}%'))
            )

    # 查询接送机已取消订单 (status=3)
    transfer_query = TransferOrder.query.filter_by(status=3)
    if date_start:
        try:
            transfer_query = transfer_query.filter(TransferOrder.cancelled_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            transfer_query = transfer_query.filter(TransferOrder.cancelled_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
        except ValueError:
            pass
    if refund_status is not None:
        transfer_query = transfer_query.filter_by(refund_status=refund_status)
    if keyword:
        transfer_query = transfer_query.filter(
            (TransferOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_phone.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_email.ilike(f'%{escape_like(keyword)}%'))
        )
    
    # 获取所有符合条件的订单
    shop_orders = (
        shop_query.order_by(ShopOrder.cancelled_at.desc().nullslast(), ShopOrder.created_at.desc()).all()
        if shop_query is not None else []
    )
    transfer_orders = transfer_query.order_by(TransferOrder.cancelled_at.desc().nullslast(), TransferOrder.created_at.desc()).all()

    # 合并并标记订单类型
    all_orders = []
    for o in shop_orders:
        order_dict = o.to_dict()
        order_dict['order_type'] = 'shop'
        order_dict['service_type'] = '商城'
        all_orders.append(order_dict)
    
    for o in transfer_orders:
        order_dict = o.to_dict()
        order_dict['order_type'] = 'transfer'
        service_type_map = {'pickup': '接机', 'dropoff': '送机', 'combo': '接送组合'}
        order_dict['service_type'] = service_type_map.get(o.service_type, o.service_type)
        all_orders.append(order_dict)
    
    # 按取消时间倒序排序
    all_orders.sort(key=lambda x: x.get('cancelled_at') or x.get('created_at') or '', reverse=True)
    
    # 手动分页
    total = len(all_orders)
    start = (page - 1) * per_page
    end = start + per_page
    page_orders = all_orders[start:end]
    
    return success_response({
        'list': page_orders,
        'total': total,
        'page': page,
        'pages': (total + per_page - 1) // per_page
    })


@api_bp.route('/admin/orders/cancelled/export', methods=['GET'])
@admin_required
def admin_export_cancelled_orders():
    """导出已取消订单为 CSV"""
    import csv
    import io
    from datetime import datetime
    from app.models import china_now
    from flask import make_response
    
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')
    refund_status = request.args.get('refund_status', type=int)
    keyword = request.args.get('keyword', '')
    
    # 不设置默认日期筛选；未选择日期时导出全部符合条件的已取消订单
    
    # 查询商城已取消订单
    shop_query = ShopOrder.query.filter_by(status=4)
    if date_start:
        try:
            shop_query = shop_query.filter(ShopOrder.cancelled_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            shop_query = shop_query.filter(ShopOrder.cancelled_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
        except ValueError:
            pass
    if refund_status is not None:
        shop_query = shop_query.filter_by(refund_status=refund_status)
    if keyword:
        shop_query = shop_query.filter(
            (ShopOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.contact_phone.ilike(f'%{escape_like(keyword)}%')) |
            (ShopOrder.contact_email.ilike(f'%{escape_like(keyword)}%'))
        )
    
    # 查询接送机已取消订单
    transfer_query = TransferOrder.query.filter_by(status=3)
    if date_start:
        try:
            transfer_query = transfer_query.filter(TransferOrder.cancelled_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            transfer_query = transfer_query.filter(TransferOrder.cancelled_at <= datetime.fromisoformat(date_end + 'T23:59:59'))
        except ValueError:
            pass
    if refund_status is not None:
        transfer_query = transfer_query.filter_by(refund_status=refund_status)
    if keyword:
        transfer_query = transfer_query.filter(
            (TransferOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_phone.ilike(f'%{escape_like(keyword)}%')) |
            (TransferOrder.contact_email.ilike(f'%{escape_like(keyword)}%'))
        )
    
    shop_orders = shop_query.order_by(ShopOrder.cancelled_at.desc().nullslast()).all()
    transfer_orders = transfer_query.order_by(TransferOrder.cancelled_at.desc().nullslast()).all()
    
    # 生成 CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 写入表头（带 BOM 以支持 Excel 正确识别 UTF-8）
    output.write('\ufeff')
    writer.writerow([
        '取消日期', '取消时间', '订单类型', '服务类型', '订单号',
        '客户姓名', '联系电话', '联系邮箱', '支付状态',
        '订单总金额', '退款金额', '退款状态', '备注'
    ])
    
    # 写入商城订单
    for o in shop_orders:
        cancelled_at = o.cancelled_at.strftime('%Y-%m-%d') if o.cancelled_at else ''
        cancelled_time = o.cancelled_at.strftime('%H:%M:%S') if o.cancelled_at else ''
        payment_status_text = '已支付' if o.payment_status == 1 else '未支付'
        refund_status_map = {0: '无需退款', 1: '待退款', 2: '已退款'}
        refund_status_text = refund_status_map.get(o.refund_status or 0, '未知')
        
        writer.writerow([
            cancelled_at,
            cancelled_time,
            '商城',
            '商城订单',
            o.order_no,
            o.contact_name,
            o.contact_phone or '',
            o.contact_email or '',
            payment_status_text,
            float(o.total_price),
            float(o.refund_amount) if o.refund_amount is not None else '',
            refund_status_text,
            o.remark or ''
        ])
    
    # 写入接送机订单
    for o in transfer_orders:
        cancelled_at = o.cancelled_at.strftime('%Y-%m-%d') if o.cancelled_at else ''
        cancelled_time = o.cancelled_at.strftime('%H:%M:%S') if o.cancelled_at else ''
        payment_status_text = '已支付' if o.payment_status == 1 else '未支付'
        refund_status_map = {0: '无需退款', 1: '待退款', 2: '已退款'}
        refund_status_text = refund_status_map.get(o.refund_status or 0, '未知')
        service_type_map = {'pickup': '接机', 'dropoff': '送机', 'combo': '接送组合'}
        service_type_text = service_type_map.get(o.service_type, o.service_type)
        
        writer.writerow([
            cancelled_at,
            cancelled_time,
            '接送机',
            service_type_text,
            o.order_no,
            o.contact_name,
            o.contact_phone or '',
            o.contact_email or '',
            payment_status_text,
            float(o.total_price),
            float(o.refund_amount) if o.refund_amount is not None else '',
            refund_status_text,
            o.remark or ''
        ])
    
    # 生成响应
    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv; charset=utf-8'
    response.headers['Content-Disposition'] = f'attachment; filename=cancelled_orders_{china_now().strftime("%Y%m%d")}.csv'
    
    return response


# ==================== 退款状态管理 ====================

@api_bp.route('/admin/orders/refund/<string:order_type>/<int:order_id>', methods=['PUT'])
@admin_required
def admin_update_refund_status(order_type, order_id):
    """更新订单退款状态（商城或接送机订单）"""
    from app.models import china_now
    
    if order_type not in ['shop', 'transfer']:
        return error_response('无效的订单类型', 400)
    
    # 查找订单
    if order_type == 'shop':
        order = ShopOrder.query.get(order_id)
    else:
        order = TransferOrder.query.get(order_id)
    
    if not order:
        return error_response('订单不存在', 404)
    
    data = request.get_json()
    new_refund_status = data.get('refund_status')
    
    if new_refund_status is None:
        return error_response('缺少退款状态参数')
    
    # 验证退款状态值
    if new_refund_status not in [0, 1, 2]:
        return error_response('无效的退款状态')
    
    old_refund_status = order.refund_status or 0
    order.refund_status = new_refund_status
    
    # 如果标记为已退款，记录退款时间和金额
    if new_refund_status == 2 and old_refund_status != 2:
        if not order.refund_time:
            order.refund_time = china_now()
        if order.refund_amount is None or order.refund_amount == 0:
            order.refund_amount = order.total_price
    
    # 如果改为无需退款，清空退款金额和时间
    if new_refund_status == 0:
        order.refund_amount = 0
        order.refund_time = None
    
    # 允许管理员修改退款金额
    if 'refund_amount' in data:
        order.refund_amount = data['refund_amount']
    
    # 允许管理员添加备注
    if 'remark' in data:
        order.remark = data['remark']
    
    db.session.commit()
    
    return success_response(order.to_dict(), '退款状态更新成功')


# ==================== 景点管理 ====================

@api_bp.route('/admin/ticket-attractions', methods=['GET'])
@admin_required
def admin_get_ticket_attractions():
    """获取景点列表（管理端）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 15, type=int)
    status = request.args.get('status', type=int)
    city = request.args.get('city')
    keyword = request.args.get('keyword', '')

    query = TicketAttraction.query
    if status is not None:
        query = query.filter_by(status=status)
    if city:
        query = query.filter_by(city=city)
    if keyword:
        query = query.filter(
            (TicketAttraction.name_zh.ilike(f'%{escape_like(keyword)}%')) |
            (TicketAttraction.name_en.ilike(f'%{escape_like(keyword)}%'))
        )

    query = query.order_by(TicketAttraction.sort_order.desc(), TicketAttraction.id.desc())
    result = paginate_query(query, page, per_page)

    return success_response({
        'list': [a.to_dict() for a in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/ticket-attractions', methods=['POST'])
@admin_required
def admin_create_ticket_attraction():
    """创建景点"""
    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'subtitle', 'desc', 'address', 'open_hours', 'visit_notice', 'refund_rule'])

    name_zh = data.get('name_zh') or data.get('name_en')
    if not name_zh:
        return error_response('景点名称不能为空')

    attraction = TicketAttraction(
        name_zh=name_zh,
        name_en=data.get('name_en'),
        name_ru=data.get('name_ru'),
        name_es=data.get('name_es'),
        subtitle_zh=data.get('subtitle_zh'), subtitle_en=data.get('subtitle_en'),
        subtitle_ru=data.get('subtitle_ru'), subtitle_es=data.get('subtitle_es'),
        desc_zh=data.get('desc_zh'), desc_en=data.get('desc_en'),
        desc_ru=data.get('desc_ru'), desc_es=data.get('desc_es'),
        address_zh=data.get('address_zh'), address_en=data.get('address_en'),
        address_ru=data.get('address_ru'), address_es=data.get('address_es'),
        open_hours_zh=data.get('open_hours_zh'), open_hours_en=data.get('open_hours_en'),
        open_hours_ru=data.get('open_hours_ru'), open_hours_es=data.get('open_hours_es'),
        visit_notice_zh=data.get('visit_notice_zh'), visit_notice_en=data.get('visit_notice_en'),
        visit_notice_ru=data.get('visit_notice_ru'), visit_notice_es=data.get('visit_notice_es'),
        refund_rule_zh=data.get('refund_rule_zh'), refund_rule_en=data.get('refund_rule_en'),
        refund_rule_ru=data.get('refund_rule_ru'), refund_rule_es=data.get('refund_rule_es'),
        cover_image=data.get('cover_image'),
        images=data.get('images', []),
        city=data.get('city'),
        category=data.get('category'),
        tags=data.get('tags', []),
        featured=data.get('featured', False),
        real_name_required=data.get('real_name_required', False),
        passport_required=data.get('passport_required', False),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    db.session.add(attraction)
    db.session.commit()

    return success_response(attraction.to_dict(), '创建成功')


@api_bp.route('/admin/ticket-attractions/<int:attraction_id>', methods=['PUT'])
@admin_required
def admin_update_ticket_attraction(attraction_id):
    """更新景点"""
    attraction = TicketAttraction.query.get(attraction_id)
    if not attraction:
        return error_response('景点不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['name', 'subtitle', 'desc', 'address', 'open_hours', 'visit_notice', 'refund_rule'])

    fields = [
        'name_zh', 'name_en', 'name_ru', 'name_es',
        'subtitle_zh', 'subtitle_en', 'subtitle_ru', 'subtitle_es',
        'desc_zh', 'desc_en', 'desc_ru', 'desc_es',
        'address_zh', 'address_en', 'address_ru', 'address_es',
        'open_hours_zh', 'open_hours_en', 'open_hours_ru', 'open_hours_es',
        'visit_notice_zh', 'visit_notice_en', 'visit_notice_ru', 'visit_notice_es',
        'refund_rule_zh', 'refund_rule_en', 'refund_rule_ru', 'refund_rule_es',
        'cover_image', 'images', 'city', 'category', 'tags',
        'featured', 'real_name_required', 'passport_required',
        'sort_order', 'status'
    ]

    for field in fields:
        if field in data:
            setattr(attraction, field, data[field])

    db.session.commit()
    return success_response(attraction.to_dict(), '更新成功')


@api_bp.route('/admin/ticket-attractions/<int:attraction_id>', methods=['DELETE'])
@admin_required
def admin_delete_ticket_attraction(attraction_id):
    """删除景点"""
    attraction = TicketAttraction.query.get(attraction_id)
    if not attraction:
        return error_response('景点不存在', 404)

    db.session.delete(attraction)
    db.session.commit()
    return success_response(None, '删除成功')


# ==================== 票种管理 ====================

@api_bp.route('/admin/ticket-packages', methods=['GET'])
@admin_required
def admin_get_ticket_packages():
    """获取票种列表（管理端）"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    attraction_id = request.args.get('attraction_id', type=int)
    ticket_type = request.args.get('ticket_type')
    status = request.args.get('status', type=int)
    keyword = request.args.get('keyword', '')

    query = TicketPackage.query
    if attraction_id:
        query = query.filter_by(attraction_id=attraction_id)
    if ticket_type:
        query = query.filter_by(ticket_type=ticket_type)
    if status is not None:
        query = query.filter_by(status=status)
    if keyword:
        query = query.filter(
            (TicketPackage.package_name_zh.ilike(f'%{escape_like(keyword)}%')) |
            (TicketPackage.package_name_en.ilike(f'%{escape_like(keyword)}%'))
        )

    query = query.order_by(TicketPackage.sort_order.desc(), TicketPackage.id.desc())
    result = paginate_query(query, page, per_page)

    return success_response({
        'list': [p.to_dict() for p in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/ticket-packages', methods=['POST'])
@admin_required
def admin_create_ticket_package():
    """创建票种"""
    data = request.get_json()
    data = auto_fill_translations(data, ['package_name', 'age_rule', 'booking_notice', 'refund_rule'])

    date_rules = data.get('date_rules')
    if date_rules and isinstance(date_rules, list):
        available_days = [item.get('date') for item in date_rules if item.get('date') and item.get('enabled', True)]
        data['available_days'] = available_days

    attraction_id = data.get('attraction_id')
    if not attraction_id:
        return error_response('请选择所属景点')

    attraction = TicketAttraction.query.get(attraction_id)
    if not attraction:
        return error_response('景点不存在', 404)

    package_name_zh = data.get('package_name_zh') or data.get('package_name_en')
    if not package_name_zh:
        return error_response('票种名称不能为空')
    if not data.get('sale_price'):
        return error_response('售价不能为空')

    pkg = TicketPackage(
        attraction_id=attraction_id,
        package_name_zh=package_name_zh,
        package_name_en=data.get('package_name_en'),
        package_name_ru=data.get('package_name_ru'),
        package_name_es=data.get('package_name_es'),
        ticket_type=data.get('ticket_type', 'adult'),
        sale_price=data.get('sale_price'),
        original_price=data.get('original_price'),
        age_rule_zh=data.get('age_rule_zh'), age_rule_en=data.get('age_rule_en'),
        age_rule_ru=data.get('age_rule_ru'), age_rule_es=data.get('age_rule_es'),
        booking_notice_zh=data.get('booking_notice_zh'), booking_notice_en=data.get('booking_notice_en'),
        booking_notice_ru=data.get('booking_notice_ru'), booking_notice_es=data.get('booking_notice_es'),
        refund_rule_zh=data.get('refund_rule_zh'), refund_rule_en=data.get('refund_rule_en'),
        refund_rule_ru=data.get('refund_rule_ru'), refund_rule_es=data.get('refund_rule_es'),
        inventory_mode=data.get('inventory_mode', 'unlimited'),
        quota_total=data.get('quota_total'),
        quota_used=0,
        available_days=data.get('available_days'),
        date_rules=data.get('date_rules'),
        sort_order=data.get('sort_order', 0),
        status=data.get('status', 1)
    )
    db.session.add(pkg)
    db.session.commit()

    return success_response(pkg.to_dict(), '创建成功')


@api_bp.route('/admin/ticket-packages/<int:package_id>', methods=['PUT'])
@admin_required
def admin_update_ticket_package(package_id):
    """更新票种"""
    pkg = TicketPackage.query.get(package_id)
    if not pkg:
        return error_response('票种不存在', 404)

    data = request.get_json()
    data = auto_fill_translations(data, ['package_name', 'age_rule', 'booking_notice', 'refund_rule'])

    date_rules = data.get('date_rules')
    if date_rules and isinstance(date_rules, list):
        available_days = [item.get('date') for item in date_rules if item.get('date') and item.get('enabled', True)]
        data['available_days'] = available_days

    fields = [
        'package_name_zh', 'package_name_en', 'package_name_ru', 'package_name_es',
        'ticket_type', 'sale_price', 'original_price',
        'age_rule_zh', 'age_rule_en', 'age_rule_ru', 'age_rule_es',
        'booking_notice_zh', 'booking_notice_en', 'booking_notice_ru', 'booking_notice_es',
        'refund_rule_zh', 'refund_rule_en', 'refund_rule_ru', 'refund_rule_es',
        'inventory_mode', 'quota_total', 'available_days', 'date_rules', 'sort_order', 'status'
    ]

    for field in fields:
        if field in data:
            setattr(pkg, field, data[field])

    db.session.commit()
    return success_response(pkg.to_dict(), '更新成功')


@api_bp.route('/admin/ticket-packages/<int:package_id>', methods=['DELETE'])
@admin_required
def admin_delete_ticket_package(package_id):
    """删除票种"""
    pkg = TicketPackage.query.get(package_id)
    if not pkg:
        return error_response('票种不存在', 404)

    db.session.delete(pkg)
    db.session.commit()
    return success_response(None, '删除成功')


# ==================== 门票加购用车价格管理 ====================

TICKET_TRANSPORT_PRICE_FIELDS = {
    'pickup_only': 'pickup_price',
    'dropoff_only': 'dropoff_price',
    'round_trip': 'round_trip_price',
    'charter': 'charter_price'
}


def _serialize_ticket_transport_price_groups(options):
    grouped = {}
    for option in options:
        key = (option.attraction_id, option.vehicle_id)
        if key not in grouped:
            grouped[key] = {
                'id': option.id,
                'attraction_id': option.attraction_id,
                'vehicle_id': option.vehicle_id,
                'vehicle': option.vehicle.to_dict() if option.vehicle else None,
                'sort_order': option.sort_order or 0,
                'status': option.status,
                'pickup_price': None,
                'dropoff_price': None,
                'round_trip_price': None,
                'charter_price': None,
                'price_ids': {}
            }
        grouped[key][TICKET_TRANSPORT_PRICE_FIELDS[option.service_type]] = float(option.price) if option.price is not None else None
        grouped[key]['price_ids'][option.service_type] = option.id
    return sorted(grouped.values(), key=lambda item: (item.get('sort_order', 0), item.get('id', 0)), reverse=True)


def _normalize_transport_group_payload(data):
    prices = {
        'pickup_only': data.get('pickup_price'),
        'dropoff_only': data.get('dropoff_price'),
        'round_trip': data.get('round_trip_price'),
        'charter': data.get('charter_price')
    }
    normalized = {}
    for service_type, raw_value in prices.items():
        if raw_value in (None, ''):
            normalized[service_type] = None
        else:
            normalized[service_type] = float(raw_value)
    return normalized


@api_bp.route('/admin/ticket-transport-pricing', methods=['GET'])
@admin_required
def admin_get_ticket_transport_pricing():
    """获取门票加购用车价格列表"""
    attraction_id = request.args.get('attraction_id', type=int)
    query = TicketTransportPrice.query
    if attraction_id:
        query = query.filter_by(attraction_id=attraction_id)

    options = query.order_by(TicketTransportPrice.sort_order.desc(), TicketTransportPrice.id.desc()).all()
    return success_response(_serialize_ticket_transport_price_groups(options))


@api_bp.route('/admin/ticket-transport-pricing', methods=['POST'])
@admin_required
def admin_create_ticket_transport_price():
    """创建门票加购用车价格"""
    data = request.get_json()

    if not data.get('attraction_id') or not data.get('vehicle_id'):
        return error_response('景点和车型不能为空')

    price_map = _normalize_transport_group_payload(data)
    if all(value is None for value in price_map.values()):
        return error_response('请至少填写一个服务价格')

    # 同一景点 + 车型 的四种价格统一覆盖保存
    existing = TicketTransportPrice.query.filter_by(
        attraction_id=data['attraction_id'],
        vehicle_id=data['vehicle_id']
    ).all()
    for item in existing:
        db.session.delete(item)
    db.session.flush()

    for service_type, price in price_map.items():
        if price is None:
            continue
        db.session.add(TicketTransportPrice(
            attraction_id=data['attraction_id'],
            vehicle_id=data['vehicle_id'],
            service_type=service_type,
            price=price,
            sort_order=data.get('sort_order', 0),
            status=data.get('status', 1)
        ))

    db.session.commit()
    saved = TicketTransportPrice.query.filter_by(
        attraction_id=data['attraction_id'],
        vehicle_id=data['vehicle_id']
    ).order_by(TicketTransportPrice.id.asc()).all()
    return success_response(_serialize_ticket_transport_price_groups(saved)[0], '创建成功')


@api_bp.route('/admin/ticket-transport-pricing/<int:price_id>', methods=['PUT'])
@admin_required
def admin_update_ticket_transport_price(price_id):
    """更新门票加购用车价格"""
    anchor = TicketTransportPrice.query.get(price_id)
    if not anchor:
        return error_response('记录不存在', 404)

    data = request.get_json()
    if not data.get('attraction_id') or not data.get('vehicle_id'):
        return error_response('景点和车型不能为空')

    price_map = _normalize_transport_group_payload(data)
    if all(value is None for value in price_map.values()):
        return error_response('请至少填写一个服务价格')

    old_rows = TicketTransportPrice.query.filter_by(
        attraction_id=anchor.attraction_id,
        vehicle_id=anchor.vehicle_id
    ).all()
    for item in old_rows:
        db.session.delete(item)

    target_rows = TicketTransportPrice.query.filter_by(
        attraction_id=data['attraction_id'],
        vehicle_id=data['vehicle_id']
    ).all()
    for item in target_rows:
        db.session.delete(item)
    db.session.flush()

    for service_type, price in price_map.items():
        if price is None:
            continue
        db.session.add(TicketTransportPrice(
            attraction_id=data['attraction_id'],
            vehicle_id=data['vehicle_id'],
            service_type=service_type,
            price=price,
            sort_order=data.get('sort_order', 0),
            status=data.get('status', 1)
        ))

    db.session.commit()
    saved = TicketTransportPrice.query.filter_by(
        attraction_id=data['attraction_id'],
        vehicle_id=data['vehicle_id']
    ).order_by(TicketTransportPrice.id.asc()).all()
    return success_response(_serialize_ticket_transport_price_groups(saved)[0], '更新成功')


@api_bp.route('/admin/ticket-transport-pricing/<int:price_id>', methods=['DELETE'])
@admin_required
def admin_delete_ticket_transport_price(price_id):
    """删除门票加购用车价格"""
    option = TicketTransportPrice.query.get(price_id)
    if not option:
        return error_response('记录不存在', 404)

    rows = TicketTransportPrice.query.filter_by(
        attraction_id=option.attraction_id,
        vehicle_id=option.vehicle_id
    ).all()
    for item in rows:
        db.session.delete(item)
    db.session.commit()
    return success_response(None, '删除成功')


# ==================== 门票订单管理 ====================

@api_bp.route('/admin/ticket-orders', methods=['GET'])
@admin_required
def admin_get_ticket_orders():
    """获取门票订单列表"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 15, type=int)
    status = request.args.get('status', type=int)
    payment_status = request.args.get('payment_status', type=int)
    keyword = request.args.get('keyword', '')
    date_start = request.args.get('date_start', '')
    date_end = request.args.get('date_end', '')

    query = TicketOrder.query

    if status is not None:
        query = query.filter_by(status=status)
    if payment_status is not None:
        query = query.filter_by(payment_status=payment_status)
    if keyword:
        query = query.filter(
            (TicketOrder.order_no.ilike(f'%{escape_like(keyword)}%')) |
            (TicketOrder.contact_name.ilike(f'%{escape_like(keyword)}%')) |
            (TicketOrder.contact_phone.ilike(f'%{escape_like(keyword)}%')) |
            (TicketOrder.contact_email.ilike(f'%{escape_like(keyword)}%')) |
            (TicketOrder.booking_no.ilike(f'%{escape_like(keyword)}%'))
        )
    if date_start:
        try:
            from datetime import datetime
            query = query.filter(TicketOrder.visit_date >= datetime.fromisoformat(date_start).date())
        except ValueError:
            pass
    if date_end:
        try:
            from datetime import datetime
            query = query.filter(TicketOrder.visit_date <= datetime.fromisoformat(date_end + 'T23:59:59').date())
        except ValueError:
            pass

    query = query.order_by(TicketOrder.created_at.desc())
    result = paginate_query(query, page, per_page)

    return success_response({
        'list': [o.to_dict() for o in result['items']],
        'total': result['total'],
        'page': result['page'],
        'pages': result['pages']
    })


@api_bp.route('/admin/ticket-orders/<int:order_id>', methods=['GET'])
@admin_required
def admin_get_ticket_order(order_id):
    """获取门票订单详情"""
    order = TicketOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    return success_response(order.to_dict())


@api_bp.route('/admin/ticket-orders/<int:order_id>/status', methods=['PUT'])
@admin_required
def admin_update_ticket_order_status(order_id):
    """更新门票订单状态"""
    from app.models import china_now

    def parse_admin_datetime(value, field_name):
        from datetime import datetime

        if value in (None, ''):
            return None
        try:
            return datetime.strptime(str(value).strip(), '%Y-%m-%d %H:%M')
        except ValueError:
            raise ValueError(f'{field_name}格式错误，请使用 YYYY-MM-DD HH:mm')

    order = TicketOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    data = request.get_json()
    new_status = data.get('status')
    if new_status is None:
        return error_response('缺少状态参数')

    if new_status not in [0, 1, 2, 3]:
        return error_response('无效的状态值')

    order.status = new_status

    if new_status == 3:  # 已取消
        order.cancelled_at = china_now()
        if order.need_transfer and not data.get('transfer_status'):
            order.transfer_status = 'cancelled'

    # 管理员备注
    if 'admin_note' in data:
        order.admin_note = data['admin_note']

    if order.need_transfer:
        try:
            if 'transfer_pickup_time' in data:
                order.transfer_pickup_time = parse_admin_datetime(data.get('transfer_pickup_time'), '接送时间')
            if 'transfer_return_time' in data:
                order.transfer_return_time = parse_admin_datetime(data.get('transfer_return_time'), '返程时间')
        except ValueError as exc:
            return error_response(str(exc))

        if 'transfer_pickup_location' in data:
            order.transfer_pickup_location = data.get('transfer_pickup_location') or None
        if 'transfer_return_location' in data:
            order.transfer_return_location = data.get('transfer_return_location') or None
        if 'transfer_admin_note' in data:
            order.transfer_admin_note = data.get('transfer_admin_note') or None
        if 'transfer_status' in data and data.get('transfer_status'):
            order.transfer_status = data.get('transfer_status')
            if data.get('transfer_status') in ['confirmed', 'scheduled', 'completed'] and not order.transfer_confirmed_at:
                order.transfer_confirmed_at = china_now()

        if isinstance(order.transfer_snapshot, dict):
            snapshot = dict(order.transfer_snapshot)
            snapshot.update({
                'pickup_time': order.transfer_pickup_time.isoformat() if order.transfer_pickup_time else None,
                'return_time': order.transfer_return_time.isoformat() if order.transfer_return_time else None,
                'pickup_location': order.transfer_pickup_location,
                'return_location': order.transfer_return_location,
                'status': order.transfer_status,
                'transfer_admin_note': order.transfer_admin_note
            })
            order.transfer_snapshot = snapshot

    db.session.commit()
    return success_response(order.to_dict(), '状态更新成功')


@api_bp.route('/admin/ticket-orders/<int:order_id>/payment', methods=['PUT'])
@admin_required
def admin_update_ticket_order_payment(order_id):
    """标记门票订单为已支付（后台确认）"""
    from app.models import china_now

    order = TicketOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    data = request.get_json()
    order.payment_status = 1
    order.payment_time = china_now()
    if 'transaction_id' in data:
        order.transaction_id = data['transaction_id']
    if 'payment_method' in data:
        order.payment_method = data['payment_method']

    db.session.commit()
    return success_response(order.to_dict(), '支付状态更新成功')


@api_bp.route('/admin/ticket-orders/<int:order_id>/voucher', methods=['POST'])
@admin_required
def admin_upload_ticket_voucher(order_id):
    """上传票据文件"""
    order = TicketOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    from flask import request as flask_request
    if 'file' not in flask_request.files:
        return error_response('没有选择文件')

    file = flask_request.files['file']
    if file.filename == '':
        return error_response('没有选择文件')

    upload_folder = current_app.config['UPLOAD_FOLDER']
    subfolder = os.path.join(upload_folder, 'tickets', 'vouchers')
    os.makedirs(subfolder, exist_ok=True)

    try:
        url = upload_file(file, subfolder)
        file_type = url.rsplit('.', 1)[-1].lower() if '.' in url else ''
        voucher = TicketVoucher(
            order_id=order_id,
            file_url=url,
            file_name=file.filename,
            file_type=file_type,
            uploaded_by=flask_request.admin_id
        )
        db.session.add(voucher)
        db.session.commit()
        return success_response(voucher.to_dict(), '票据上传成功')
    except Exception as e:
        current_app.logger.error(f'票据上传失败: {e}')
        return error_response(f'上传失败: {str(e)}', 500)


@api_bp.route('/admin/ticket-orders/<int:order_id>/voucher', methods=['DELETE'])
@admin_required
def admin_delete_ticket_voucher(order_id):
    """删除票据文件"""
    voucher_id = request.args.get('voucher_id', type=int)
    if not voucher_id:
        return error_response('缺少票据ID')

    voucher = TicketVoucher.query.get(voucher_id)
    if not voucher or voucher.order_id != order_id:
        return error_response('票据不存在', 404)

    db.session.delete(voucher)
    db.session.commit()
    return success_response(None, '删除成功')


@api_bp.route('/admin/ticket-orders/<int:order_id>/send-voucher', methods=['POST'])
@admin_required
def admin_send_ticket_voucher(order_id):
    """标记票据已发送给客户"""
    from app.models import china_now

    order = TicketOrder.query.get(order_id)
    if not order:
        return error_response('订单不存在', 404)

    order.voucher_delivery_status = 1
    for v in order.vouchers:
        if not v.sent_to_customer:
            v.sent_to_customer = True
            v.sent_at = china_now()

    db.session.commit()
    return success_response(order.to_dict(), '已标记为已发送')
