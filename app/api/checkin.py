"""境外人员住宿登记 API
公开端：上传护照照片（私密存储）+ 提交登记
管理端：列表 / 状态更新 / 删除 / Excel 导出 / 护照图片鉴权代理
"""
import re
from datetime import datetime, date, timedelta
from flask import request, current_app, Response
from app.api import api_bp
from app.models import GuestRegistration, china_now
from app import db
from app.utils import success_response, error_response, admin_required, paginate_query, escape_like
from app.utils.storage import upload_private_file, get_private_file

VALID_PLATFORMS = {'booking', 'trip', 'agoda', 'expedia'}
VALID_DOC_TYPES = {'passport', 'hkmo', 'taiwan'}


def _clean_booking_no(value):
    """清洗预约单号：去掉小数点和所有空白字符。
    Booking.com 的确认号显示为 6231.516.538 这种带点格式，客户直接复制会带点。"""
    return re.sub(r'[.\s ]+', '', value or '')


def _private_folder():
    return current_app.config.get('PRIVATE_UPLOAD_FOLDER')


# ==================== 公开接口 ====================

@api_bp.route('/guest-registrations/upload', methods=['POST'])
def guest_registration_upload():
    """上传护照/手持照片（私密存储，仅返回 key）"""
    if 'file' not in request.files:
        return error_response('No file selected')
    file = request.files['file']
    if not file.filename:
        return error_response('No file selected')

    # 额外允许 HEIC/HEIF（iPhone 默认拍照格式），服务端自动转为 JPEG
    allowed = set(current_app.config.get('ALLOWED_EXTENSIONS', {'png', 'jpg', 'jpeg', 'gif', 'webp'})) | {'heic', 'heif'}
    if '.' not in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed:
        return error_response('Unsupported file type. Please upload a JPG, PNG or HEIC photo')

    try:
        key = upload_private_file(file, _private_folder())
        return success_response({'key': key}, 'Uploaded')
    except ValueError as e:
        if 'HEIC_NOT_SUPPORTED' in str(e):
            return error_response('This photo format (HEIC) could not be processed. Please change your camera format to "Most Compatible" in iPhone Settings > Camera > Formats, or take a screenshot of the photo and upload that instead')
        current_app.logger.error(f'护照照片上传失败: {e}')
        return error_response('Upload failed, please try again', 500)
    except Exception as e:
        current_app.logger.error(f'护照照片上传失败: {e}')
        return error_response('Upload failed, please try again', 500)


@api_bp.route('/guest-registrations', methods=['POST'])
def create_guest_registration():
    """提交住宿登记"""
    data = request.get_json() or {}

    platform = (data.get('platform') or '').strip().lower()
    if platform not in VALID_PLATFORMS:
        return error_response('Invalid booking platform')

    booking_no = _clean_booking_no(data.get('booking_no'))
    surname = (data.get('surname') or '').strip()
    given_name = (data.get('given_name') or '').strip()
    middle_name = (data.get('middle_name') or '').strip() or None
    if not booking_no:
        return error_response('Booking number is required')
    if not surname or not given_name:
        return error_response('Name is required')
    if len(booking_no) > 100 or len(surname) > 100 or len(given_name) > 100:
        return error_response('Input too long')

    dob_raw = (data.get('date_of_birth') or '').strip()
    try:
        dob = date.fromisoformat(dob_raw)
    except (ValueError, TypeError):
        return error_response('Invalid date of birth')
    if dob > date.today() or dob.year < 1900:
        return error_response('Invalid date of birth')

    passport_image = (data.get('passport_image') or '').strip()
    handheld_image = (data.get('handheld_image') or '').strip()
    if not passport_image.startswith('private/') or not handheld_image.startswith('private/'):
        return error_response('Please upload both passport photos')

    document_type = (data.get('document_type') or 'passport').strip().lower()
    if document_type not in VALID_DOC_TYPES:
        document_type = 'passport'

    # 证件号归一化：去空白、转大写，便于判重
    document_no = re.sub(r'\s+', '', data.get('document_no') or '').upper()
    if not document_no:
        return error_response('Document number is required')
    if len(document_no) > 50:
        return error_response('Document number too long')

    def _parse_date(raw):
        try:
            return date.fromisoformat((raw or '').strip())
        except (ValueError, TypeError):
            return None

    checkin_date = _parse_date(data.get('checkin_date'))
    checkout_date = _parse_date(data.get('checkout_date'))
    if not checkin_date or not checkout_date:
        return error_response('Please select check-in and check-out dates')
    if checkout_date <= checkin_date:
        return error_response('Check-out date must be after check-in date')

    # 重复登记过滤：同一证件号在同一个订单里只登记一次
    # （不做全局证件号判重——老客户之后用新订单再来住是合法场景）
    existing = GuestRegistration.query.filter(
        db.func.upper(GuestRegistration.document_no) == document_no,
        GuestRegistration.platform == platform,
        GuestRegistration.booking_no == booking_no,
    ).first()
    if existing:
        return success_response({'id': existing.id, 'duplicate': True}, 'Already registered')

    reg = GuestRegistration(
        platform=platform,
        booking_no=booking_no,
        document_type=document_type,
        document_no=document_no,
        checkin_date=checkin_date,
        checkout_date=checkout_date,
        surname=surname,
        given_name=given_name,
        middle_name=middle_name,
        date_of_birth=dob,
        passport_image=passport_image,
        handheld_image=handheld_image,
        lang=(data.get('lang') or 'en')[:10],
        status=0,
    )
    db.session.add(reg)
    db.session.commit()
    return success_response({'id': reg.id}, 'Registration successful')


# ==================== 管理端接口 ====================

@api_bp.route('/admin/guest-registrations', methods=['GET'])
@admin_required
def admin_list_guest_registrations():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 15))
    keyword = (request.args.get('keyword') or '').strip()
    status = request.args.get('status', '')
    date_start = (request.args.get('date_start') or '').strip()
    date_end = (request.args.get('date_end') or '').strip()

    checkin_filter = (request.args.get('checkin_date') or '').strip()

    query = GuestRegistration.query
    if keyword:
        like = f'%{escape_like(keyword)}%'
        query = query.filter(
            db.or_(
                GuestRegistration.surname.ilike(like),
                GuestRegistration.given_name.ilike(like),
                GuestRegistration.booking_no.ilike(like),
                GuestRegistration.document_no.ilike(like),
                GuestRegistration.room_note.ilike(like),
            )
        )
    if status != '':
        try:
            query = query.filter(GuestRegistration.status == int(status))
        except (ValueError, TypeError):
            pass
    # 按入住日期筛选（每天查"今天入住"的客人）
    if checkin_filter:
        try:
            query = query.filter(GuestRegistration.checkin_date == date.fromisoformat(checkin_filter))
        except ValueError:
            pass
    if date_start:
        try:
            query = query.filter(GuestRegistration.created_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            query = query.filter(GuestRegistration.created_at < datetime.fromisoformat(date_end) + timedelta(days=1))
        except ValueError:
            pass

    query = query.order_by(GuestRegistration.created_at.desc())
    result = paginate_query(query, page=page, per_page=per_page)
    return success_response({
        'list': [r.to_dict() for r in result['items']],
        'total': result['total'],
        'page': result['page'],
        'per_page': result['per_page'],
        'pages': result['pages'],
    })


def _group_key(reg):
    """分组键：手动合并的 group_id 优先，否则按 平台+预约单号 自动分组。
    单号做归一化（去点/去空白/转小写），避免复制粘贴带来的格式差异导致同单号分成多组。"""
    if reg.group_id:
        return reg.group_id
    return f'bn:{reg.platform}:{_clean_booking_no(reg.booking_no).lower()}'


@api_bp.route('/admin/guest-registrations/grouped', methods=['GET'])
@admin_required
def admin_grouped_guest_registrations():
    """按订单分组的登记列表（一组 = 一个房间/订单的多位客人）"""
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 10))
    keyword = (request.args.get('keyword') or '').strip()
    checkin_filter = (request.args.get('checkin_date') or '').strip()
    date_start = (request.args.get('date_start') or '').strip()
    date_end = (request.args.get('date_end') or '').strip()

    query = GuestRegistration.query
    if keyword:
        like = f'%{escape_like(keyword)}%'
        query = query.filter(
            db.or_(
                GuestRegistration.surname.ilike(like),
                GuestRegistration.given_name.ilike(like),
                GuestRegistration.booking_no.ilike(like),
                GuestRegistration.document_no.ilike(like),
                GuestRegistration.room_note.ilike(like),
            )
        )
    # 按入住日期筛选（每天查"今天入住"的客人）
    if checkin_filter:
        try:
            query = query.filter(GuestRegistration.checkin_date == date.fromisoformat(checkin_filter))
        except ValueError:
            pass
    if date_start:
        try:
            query = query.filter(GuestRegistration.created_at >= datetime.fromisoformat(date_start))
        except ValueError:
            pass
    if date_end:
        try:
            query = query.filter(GuestRegistration.created_at < datetime.fromisoformat(date_end) + timedelta(days=1))
        except ValueError:
            pass

    rows = query.order_by(GuestRegistration.created_at.desc()).all()

    # 关键词命中组内任意成员时，整组完整显示（重新拉全组成员）
    if keyword and rows:
        keys = {_group_key(r) for r in rows}
        all_rows = GuestRegistration.query.order_by(GuestRegistration.created_at.desc()).all()
        rows = [r for r in all_rows if _group_key(r) in keys]

    groups = {}
    order = []
    for r in rows:
        key = _group_key(r)
        if key not in groups:
            groups[key] = {
                'key': key,
                'platform': r.platform,
                'booking_no': r.booking_no,
                'room_note': None,
                'count': 0,
                'declared_count': 0,
                'checkin_date': r.checkin_date.isoformat() if r.checkin_date else None,
                'checkout_date': r.checkout_date.isoformat() if r.checkout_date else None,
                'latest_at': r.created_at.isoformat() if r.created_at else None,
                'items': [],
            }
            order.append(key)
        g = groups[key]
        g['count'] += 1
        if r.status == 1:
            g['declared_count'] += 1
        if not g['room_note'] and r.room_note:
            g['room_note'] = r.room_note
        g['items'].append(r.to_dict())

    group_list = [groups[k] for k in order]
    total = len(group_list)
    start_idx = (page - 1) * per_page
    paged = group_list[start_idx:start_idx + per_page]

    return success_response({
        'list': paged,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page if per_page else 1,
    })


@api_bp.route('/admin/guest-registrations/batch-status', methods=['PUT'])
@admin_required
def admin_batch_guest_registration_status():
    """批量更新申报状态（整组一键标记）"""
    data = request.get_json() or {}
    ids = data.get('ids') or []
    new_status = data.get('status')
    if not ids or new_status not in (0, 1):
        return error_response('参数错误')
    regs = GuestRegistration.query.filter(GuestRegistration.id.in_(ids)).all()
    for r in regs:
        r.status = new_status
    db.session.commit()
    return success_response({'updated': len(regs)}, f'已更新 {len(regs)} 条记录')


@api_bp.route('/admin/guest-registrations/room-note', methods=['PUT'])
@admin_required
def admin_set_guest_registration_room_note():
    """设置房间备注（同步到组内所有记录）"""
    data = request.get_json() or {}
    ids = data.get('ids') or []
    room_note = (data.get('room_note') or '').strip()[:100] or None
    if not ids:
        return error_response('参数错误')
    regs = GuestRegistration.query.filter(GuestRegistration.id.in_(ids)).all()
    for r in regs:
        r.room_note = room_note
    db.session.commit()
    return success_response({'updated': len(regs)}, '房间备注已保存')


@api_bp.route('/admin/guest-registrations/merge', methods=['POST'])
@admin_required
def admin_merge_guest_registrations():
    """手动合并多条登记记录为一组（处理单号输错等情况）"""
    import uuid
    data = request.get_json() or {}
    ids = data.get('ids') or []
    if len(ids) < 2:
        return error_response('请至少选择两条记录')
    regs = GuestRegistration.query.filter(GuestRegistration.id.in_(ids)).all()
    if len(regs) < 2:
        return error_response('记录不存在')
    new_group = f'mg-{uuid.uuid4().hex[:16]}'
    # 备注取组内第一个非空的
    note = next((r.room_note for r in regs if r.room_note), None)
    for r in regs:
        r.group_id = new_group
        if note and not r.room_note:
            r.room_note = note
    db.session.commit()
    return success_response({'group_id': new_group, 'merged': len(regs)}, f'已合并 {len(regs)} 条记录')


@api_bp.route('/admin/guest-registrations/<int:reg_id>/ungroup', methods=['POST'])
@admin_required
def admin_ungroup_guest_registration(reg_id):
    """把某条记录移出当前分组（变成独立一组）"""
    reg = GuestRegistration.query.get(reg_id)
    if not reg:
        return error_response('登记记录不存在', 404)
    reg.group_id = f'solo-{reg.id}'
    db.session.commit()
    return success_response(reg.to_dict(), '已移出该组')


@api_bp.route('/admin/guest-registrations/<int:reg_id>/status', methods=['PUT'])
@admin_required
def admin_update_guest_registration_status(reg_id):
    reg = GuestRegistration.query.get(reg_id)
    if not reg:
        return error_response('登记记录不存在', 404)
    data = request.get_json() or {}
    new_status = data.get('status')
    if new_status not in (0, 1):
        return error_response('无效的状态值')
    reg.status = new_status
    db.session.commit()
    return success_response(reg.to_dict(), '状态已更新')


@api_bp.route('/admin/guest-registrations/<int:reg_id>', methods=['DELETE'])
@admin_required
def admin_delete_guest_registration(reg_id):
    reg = GuestRegistration.query.get(reg_id)
    if not reg:
        return error_response('登记记录不存在', 404)
    db.session.delete(reg)
    db.session.commit()
    return success_response(None, '删除成功')


@api_bp.route('/admin/guest-doc/<path:key>', methods=['GET'])
@admin_required
def admin_get_guest_doc(key):
    """护照照片鉴权代理（仅管理端可访问，支持 ?token= 参数）"""
    data, content_type = get_private_file(key, _private_folder())
    if data is None:
        return error_response('图片不存在', 404)
    resp = Response(data, mimetype=content_type or 'image/jpeg')
    resp.headers['Cache-Control'] = 'private, max-age=300'
    return resp


@api_bp.route('/admin/guest-registrations/export', methods=['GET'])
@admin_required
def admin_export_guest_registrations():
    """导出住宿登记 Excel。默认最近一个月，支持 date_start/date_end 自选范围。"""
    from io import BytesIO
    from flask import send_file

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error_response('服务器缺少 openpyxl 依赖，请联系运维', 500)

    now = china_now()
    checkin_filter = (request.args.get('checkin_date') or '').strip()
    date_start = (request.args.get('date_start') or '').strip()
    date_end = (request.args.get('date_end') or '').strip()

    if checkin_filter:
        # 按入住日期导出当天入住的客人
        try:
            target_day = date.fromisoformat(checkin_filter)
        except ValueError:
            return error_response('入住日期格式错误')
        rows = GuestRegistration.query.filter(
            GuestRegistration.checkin_date == target_day
        ).order_by(GuestRegistration.created_at.desc()).all()
        start = end = None
    else:
        if date_start:
            try:
                start = datetime.fromisoformat(date_start)
            except ValueError:
                start = now - timedelta(days=30)
        else:
            start = now - timedelta(days=30)
        if date_end:
            try:
                end = datetime.fromisoformat(date_end) + timedelta(days=1)
            except ValueError:
                end = now
        else:
            end = now

        rows = GuestRegistration.query.filter(
            GuestRegistration.created_at >= start,
            GuestRegistration.created_at <= end,
        ).order_by(GuestRegistration.created_at.desc()).all()

    # 按组排序：同一房间/订单的客人相邻排列（保持组的最新登记时间倒序）
    group_sizes = {}
    group_first_seen = {}
    for idx, r in enumerate(rows):
        key = _group_key(r)
        group_sizes[key] = group_sizes.get(key, 0) + 1
        if key not in group_first_seen:
            group_first_seen[key] = idx
    rows.sort(key=lambda r: (group_first_seen[_group_key(r)], r.id))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '住宿登记'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='4A3728')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='DDD0BC')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    headers = ['ID', '姓 (Surname)', '名 (Given Name)', '中间名', '出生日期', '证件类型', '证件号码', '预订平台', '预约单号', '房间备注', '同组人数', '入住日期', '离开日期', '申报状态', '登记时间']
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    platform_labels = {'booking': 'Booking.com', 'trip': 'Trip.com', 'agoda': 'Agoda', 'expedia': 'Expedia'}
    doc_labels = {'passport': '外国人护照', 'hkmo': '港澳通行证', 'taiwan': '台湾通行证'}
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r.id)
        ws.cell(row=i, column=2, value=r.surname)
        ws.cell(row=i, column=3, value=r.given_name)
        ws.cell(row=i, column=4, value=r.middle_name or '')
        ws.cell(row=i, column=5, value=r.date_of_birth.isoformat() if r.date_of_birth else '')
        ws.cell(row=i, column=6, value=doc_labels.get(r.document_type or 'passport', r.document_type))
        ws.cell(row=i, column=7, value=r.document_no or '')
        ws.cell(row=i, column=8, value=platform_labels.get(r.platform, r.platform))
        ws.cell(row=i, column=9, value=r.booking_no)
        ws.cell(row=i, column=10, value=r.room_note or '')
        ws.cell(row=i, column=11, value=group_sizes.get(_group_key(r), 1))
        ws.cell(row=i, column=12, value=r.checkin_date.isoformat() if r.checkin_date else '')
        ws.cell(row=i, column=13, value=r.checkout_date.isoformat() if r.checkout_date else '')
        ws.cell(row=i, column=14, value='已申报' if r.status == 1 else '待申报')
        ws.cell(row=i, column=15, value=r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '')

    widths = [8, 16, 16, 14, 14, 14, 20, 14, 24, 16, 10, 13, 13, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    if checkin_filter:
        filename = f'住宿登记_入住{checkin_filter.replace("-", "")}.xlsx'
    else:
        filename = f'住宿登记_{start.strftime("%Y%m%d")}-{(end - timedelta(days=1)).strftime("%Y%m%d") if date_end else now.strftime("%Y%m%d")}.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
